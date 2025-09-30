# ============================================
# Chat-bot v2 — SINGLE FILE (bot.py)
# Version: v8.4 (2025-09-29)
# База: стабильная v8.3.3
#
# Что добавлено по запросу:
# 1) Отклонения — комментарии строго РЕПЛАЕМ на спец-сообщение, бот явно об этом говорит.
# 2) «Отменить отклонение» (лидер): комментарий реплаем; заявка возвращается исполнителю и закрепляется за ним.
# 3) «Нет доступа к помещению» — флоу аналогичен «Не к УТО» (через лидера).
# 4) Уточнения: автор отвечает реплаем, ответ уходит исполнителю в чат заявки; снова доступны кнопки.
# 5) Excel-лог: даты исправлены; добавлено поле initial_group (первая группа маршрутизации).
# ============================================

# asdasd

import os
import io
import csv
import json
import time
import uuid
import sqlite3
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple, Set
from datetime import datetime, UTC, timedelta

from dotenv import load_dotenv
from loguru import logger
from rapidfuzz import fuzz
from html import escape as html_escape

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    InputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.error import TelegramError
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

# ============================================
# ПУТИ / ЛОГИ / ENV
# ============================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent  # ...\Chat-bot
DATA_DIR = PROJECT_ROOT / "data"
LOGS_DIR = PROJECT_ROOT / "logs"
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = DATA_DIR / "bot.db"

def setup_logging(logs_dir: Path) -> None:
    logs_dir.mkdir(parents=True, exist_ok=True)
    logger.remove()
    logger.add(
        sink=lambda m: print(m, end=""),
        format="<green>{time:YYYY-MM-DD HH:mm:ss.SSS}</green> | <level>{level}</level> | {message}",
        level="INFO",
    )
    logger.add(
        logs_dir / "bot.log",
        rotation="5 MB",
        retention="7 days",
        compression="zip",
        encoding="utf-8",
        format="{time:YYYY-MM-DD HH:mm:ss.SSS} | {level} | {message}",
        level="INFO",
    )

def load_env(project_root: Path) -> None:
    env_path = project_root / ".env"
    if env_path.exists():
        load_dotenv(env_path)
    else:
        load_dotenv()

# ============================================
# JSONL ПЕРСИСТ (оставляем)
# ============================================

TICKETS_FILE = DATA_DIR / "tickets.jsonl"
FEEDBACK_FILE = DATA_DIR / "feedback.jsonl"

def _append_jsonl(path: Path, record: Dict[str, Any]) -> None:
    record = {**record, "ts": record.get("ts") or datetime.now(UTC).isoformat(timespec="seconds")}
    with path.open("a", encoding="utf-8") as f:
        json.dump(record, f, ensure_ascii=False)
        f.write("\n")

def save_ticket_event_jsonl(event: Dict[str, Any]) -> None:
    _append_jsonl(TICKETS_FILE, event)

def save_feedback_jsonl(event: Dict[str, Any]) -> None:
    _append_jsonl(FEEDBACK_FILE, event)

# ============================================
# SQLITE: СХЕМА / МИГРАЦИИ / УТИЛИТЫ
# ============================================

SQL_SCHEMA = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS users (
    telegram_user_id INTEGER PRIMARY KEY,
    phone_e164 TEXT NOT NULL,
    full_name TEXT,
    roles TEXT NOT NULL,
    verified_at TEXT NOT NULL,
    active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS tickets (
    ticket_id TEXT PRIMARY KEY,
    author_id INTEGER,
    author_name TEXT,
    text TEXT,
    group_name TEXT,           -- текущая группа (на данный момент)
    initial_group TEXT,        -- ПЕРВАЯ группа, куда отправили заявку
    category TEXT,
    created_ts TEXT,
    queued_ts TEXT,
    accepted_ts TEXT,
    rejected_ts TEXT,
    closed_ts TEXT,
    final_status TEXT,
    executor_id INTEGER,
    executor_name TEXT,
    reject_reason_code TEXT,
    reject_comment TEXT,
    leader_id INTEGER,
    leader_name TEXT,
    leader_decision_ts TEXT,
    rerouted_to_group TEXT,
    rerouted_ts TEXT,
    clarify_question TEXT,
    clarify_requested_ts TEXT,
    clarify_answer TEXT,
    clarify_answered_ts TEXT,
    group_chat_id INTEGER,
    group_message_id INTEGER,
    updated_ts TEXT
);

CREATE TABLE IF NOT EXISTS ticket_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id TEXT NOT NULL,
    event TEXT NOT NULL,
    ts_utc TEXT NOT NULL,
    author_id INTEGER,
    executor_id INTEGER,
    group_name TEXT,
    category TEXT,
    payload_json TEXT
);

CREATE TABLE IF NOT EXISTS sync_state (
    system TEXT PRIMARY KEY,
    last_export_ts TEXT
);

CREATE INDEX IF NOT EXISTS idx_events_ticket_ts ON ticket_events(ticket_id, ts_utc);
CREATE INDEX IF NOT EXISTS idx_tickets_updated ON tickets(updated_ts);
"""

TICKETS_EXPECTED_COLS = [
    "ticket_id","author_id","author_name","text","group_name","initial_group","category",
    "created_ts","queued_ts","accepted_ts","rejected_ts","closed_ts",
    "final_status","executor_id","executor_name","reject_reason_code","reject_comment",
    "leader_id","leader_name","leader_decision_ts","rerouted_to_group","rerouted_ts",
    "clarify_question","clarify_requested_ts","clarify_answer","clarify_answered_ts",
    "group_chat_id","group_message_id","updated_ts"
]
USERS_EXPECTED_COLS = [
    "telegram_user_id","phone_e164","full_name","roles","verified_at","active"
]
EVENTS_EXPECTED_COLS = [
    "id","ticket_id","event","ts_utc","author_id","executor_id","group_name","category","payload_json"
]
SYNC_EXPECTED_COLS = [
    "system","last_export_ts"
]

def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def _ensure_columns(conn: sqlite3.Connection, table: str, expected_cols: List[str]) -> List[str]:
    cur = conn.execute(f"PRAGMA table_info({table})")
    existing = [row["name"] for row in cur.fetchall()]
    missing = [c for c in expected_cols if c not in existing]
    for name in missing:
        logger.warning(f"[DB MIGRATION] Adding missing column {table}.{name} TEXT")
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {name} TEXT")
    if missing:
        logger.info(f"[DB MIGRATION] {table}: added {missing}")
    cur2 = conn.execute(f"PRAGMA table_info({table})")
    existing2 = [row["name"] for row in cur2.fetchall()]
    return existing2

def db_init() -> None:
    with db() as conn:
        conn.executescript(SQL_SCHEMA)
        users_cols   = _ensure_columns(conn, "users", USERS_EXPECTED_COLS)
        tickets_cols = _ensure_columns(conn, "tickets", TICKETS_EXPECTED_COLS)
        events_cols  = _ensure_columns(conn, "ticket_events", EVENTS_EXPECTED_COLS)
        sync_cols    = _ensure_columns(conn, "sync_state", SYNC_EXPECTED_COLS)
        logger.info(f"[DB SCHEMA] users   = {users_cols}")
        logger.info(f"[DB SCHEMA] tickets = {tickets_cols}")
        logger.info(f"[DB SCHEMA] events  = {events_cols}")
        logger.info(f"[DB SCHEMA] sync    = {sync_cols}")

def iso_now() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")

def _get_table_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.execute(f"PRAGMA table_info({table})")
    return [row["name"] for row in cur.fetchall()]

def _dict_select(d: Dict[str, Any], keys: List[str]) -> Dict[str, Any]:
    return {k: d.get(k) for k in keys}

def _dynamic_insert(conn: sqlite3.Connection, table: str, row: Dict[str, Any], preferred_order: List[str]) -> None:
    actual_cols = _get_table_columns(conn, table)
    cols = [c for c in preferred_order if c in actual_cols]
    vals = [_dict_select(row, cols).get(c) for c in cols]
    placeholders = ",".join("?" for _ in cols)
    col_list = ",".join(cols)
    sql = f"INSERT INTO {table}({col_list}) VALUES({placeholders})"
    logger.warning(f"[DB DYNAMIC INSERT] {table}: using cols={cols}")
    conn.execute(sql, tuple(vals))

def _dynamic_update(conn: sqlite3.Connection, table: str, row: Dict[str, Any], where_key: str, preferred_update_cols: List[str]) -> None:
    actual_cols = _get_table_columns(conn, table)
    # Берём только те колонки, которые реально присутствуют и для которых передано НЕ None
    cols = [c for c in preferred_update_cols if c in actual_cols and c != where_key and (row.get(c) is not None)]
    if not cols:
        # обновлять нечего — но чтобы пометить "updated_ts", можно отдельно коснуться этого поля, если оно задано
        if (row.get("updated_ts") is not None) and ("updated_ts" in actual_cols):
            conn.execute(f"UPDATE {table} SET updated_ts=? WHERE {where_key}=?", (row["updated_ts"], row[where_key]))
        return
    set_expr = ",".join(f"{c}=?" for c in cols)
    vals = [row.get(c) for c in cols]
    sql = f"UPDATE {table} SET {set_expr} WHERE {where_key}=?"
    logger.warning(f"[DB DYNAMIC UPDATE] {table}: using cols(non-null)={cols}")
    conn.execute(sql, tuple(vals + [row[where_key]]))

def db_upsert_user(telegram_user_id: int, phone: str, full_name: str, roles_csv: str) -> None:
    with db() as conn:
        conn.execute("""
            INSERT INTO users(telegram_user_id, phone_e164, full_name, roles, verified_at, active)
            VALUES (?, ?, ?, ?, ?, 1)
            ON CONFLICT(telegram_user_id) DO UPDATE SET
                phone_e164=excluded.phone_e164,
                full_name=excluded.full_name,
                roles=excluded.roles,
                verified_at=excluded.verified_at,
                active=1
        """, (telegram_user_id, phone, full_name, roles_csv, iso_now()))

def db_get_user_roles(telegram_user_id: int) -> Set[str]:
    with db() as conn:
        r = conn.execute("SELECT roles, active FROM users WHERE telegram_user_id=?", (telegram_user_id,)).fetchone()
        if not r or r["active"] != 1:
            return set()
        roles_csv = r["roles"] or ""
        return {x.strip() for x in roles_csv.split(",") if x.strip()}

def db_find_users_by_role_prefix(prefix: str) -> List[sqlite3.Row]:
    with db() as conn:
        cur = conn.execute("SELECT * FROM users WHERE active=1 AND roles LIKE ?", (f"%{prefix}%",))
        return cur.fetchall()

def db_insert_event(ev: Dict[str, Any]) -> None:
    with db() as conn:
        conn.execute("""
            INSERT INTO ticket_events(ticket_id, event, ts_utc, author_id, executor_id, group_name, category, payload_json)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ev.get("ticket_id") or ev.get("id"),
            ev["event"],
            ev.get("ts") or iso_now(),
            ev.get("submitter_id"),
            ev.get("executor_id"),
            (ev.get("classification") or {}).get("group") or ev.get("group"),
            (ev.get("classification") or {}).get("category") or ev.get("category"),
            json.dumps(ev, ensure_ascii=False)
        ))

def db_upsert_ticket_snapshot(t: Dict[str, Any]) -> None:
    now = iso_now()
    row = {
        "ticket_id": t["id"],
        "author_id": t.get("submitter_id"),
        "author_name": t.get("submitter_name"),
        "text": t.get("text"),
        "group_name": (t.get("classification") or {}).get("group"),
        "initial_group": t.get("initial_group"),
        "category": (t.get("classification") or {}).get("category"),
        "created_ts": t.get("created_ts"),
        "queued_ts": t.get("queued_ts"),
        "accepted_ts": t.get("accepted_ts"),
        "rejected_ts": t.get("rejected_ts"),
        "closed_ts": t.get("closed_ts"),
        "final_status": t.get("status"),
        "executor_id": t.get("executor_id"),
        "executor_name": t.get("executor_name"),
        "reject_reason_code": t.get("reject_reason_code"),
        "reject_comment": t.get("reject_comment"),
        "leader_id": t.get("leader_id"),
        "leader_name": t.get("leader_name"),
        "leader_decision_ts": t.get("leader_decision_ts"),
        "rerouted_to_group": t.get("rerouted_to_group"),
        "rerouted_ts": t.get("rerouted_ts"),
        "clarify_question": t.get("clarify_question"),
        "clarify_requested_ts": t.get("clarify_requested_ts"),
        "clarify_answer": t.get("clarify_answer"),
        "clarify_answered_ts": t.get("clarify_answered_ts"),
        "group_chat_id": t.get("group_chat_id"),
        "group_message_id": t.get("group_message_id"),
        "updated_ts": now,
    }
    with db() as conn:
        _ensure_columns(conn, "tickets", TICKETS_EXPECTED_COLS)
        existed = conn.execute("SELECT 1 FROM tickets WHERE ticket_id=?", (t["id"],)).fetchone()
        try:
            if existed:
                _dynamic_update(conn, "tickets", row, "ticket_id", TICKETS_EXPECTED_COLS)
            else:
                _dynamic_insert(conn, "tickets", row, TICKETS_EXPECTED_COLS)
        except sqlite3.OperationalError as e:
            logger.error(f"[DB FALLBACK] tickets upsert failed: {e}")
            _dynamic_update(conn, "tickets", row, "ticket_id", TICKETS_EXPECTED_COLS)

def db_touch_ticket_timestamp(ticket_id: str, field: str, ts: Optional[str] = None) -> None:
    val = ts or iso_now()
    with db() as conn:
        if field not in _get_table_columns(conn, "tickets"):
            logger.warning(f"[DB MIGRATION] Adding missing column tickets.{field} TEXT (touch)")
            conn.execute(f"ALTER TABLE tickets ADD COLUMN {field} TEXT")
        conn.execute(f"UPDATE tickets SET {field}=?, updated_ts=? WHERE ticket_id=?", (val, iso_now(), ticket_id))

def db_fetch_tickets_rows() -> List[Dict[str, Any]]:
    with db() as conn:
        cur = conn.execute("""
            SELECT ticket_id, initial_group, group_name AS "group", category,
                   author_id, author_name, executor_id, executor_name,
                   created_ts, queued_ts, accepted_ts, rejected_ts, closed_ts,
                   final_status, reject_reason_code, reject_comment, leader_name, rerouted_to_group,
                   clarify_question, clarify_requested_ts, clarify_answer, clarify_answered_ts
            FROM tickets ORDER BY COALESCE(created_ts, updated_ts) ASC
        """)
        return [dict(r) for r in cur.fetchall()]

def db_update_from_events() -> None:
    with db() as conn:
        cur = conn.execute("""
            SELECT ticket_id,
                   MIN(CASE WHEN event='new_text' THEN ts_utc END) AS created_ts,
                   MIN(CASE WHEN event='queued_to_group' THEN ts_utc END) AS queued_ts,
                   MIN(CASE WHEN event='accepted' THEN ts_utc END) AS accepted_ts,
                   MIN(CASE WHEN event='rejected' THEN ts_utc END) AS rejected_ts,
                   MIN(CASE WHEN event='closed_by_executor' THEN ts_utc END) AS closed_ts
            FROM ticket_events
            GROUP BY ticket_id
        """)
        rows = cur.fetchall()
        for r in rows:
            final_status = None
            if r["closed_ts"]:
                final_status = "closed"
            elif r["rejected_ts"]:
                final_status = "rejected"
            elif r["accepted_ts"]:
                final_status = "accepted"
            elif r["queued_ts"]:
                final_status = "queued"
            elif r["created_ts"]:
                final_status = "created"
            if final_status:
                conn.execute("""
                    UPDATE tickets SET created_ts=COALESCE(created_ts, ?),
                                       queued_ts=COALESCE(queued_ts, ?),
                                       accepted_ts=COALESCE(accepted_ts, ?),
                                       rejected_ts=COALESCE(rejected_ts, ?),
                                       closed_ts=COALESCE(closed_ts, ?),
                                       final_status=?,
                                       updated_ts=?
                    WHERE ticket_id=?
                """, (r["created_ts"], r["queued_ts"], r["accepted_ts"], r["rejected_ts"], r["closed_ts"],
                      final_status, iso_now(), r["ticket_id"]))

def db_get_last_export_ts(system: str) -> Optional[str]:
    with db() as conn:
        r = conn.execute("SELECT last_export_ts FROM sync_state WHERE system=?", (system,)).fetchone()
        return r["last_export_ts"] if r and r["last_export_ts"] else None

def db_set_last_export_ts(system: str, ts: str) -> None:
    with db() as conn:
        conn.execute("""
            INSERT INTO sync_state(system, last_export_ts) VALUES(?, ?)
            ON CONFLICT(system) DO UPDATE SET last_export_ts=excluded.last_export_ts
        """, (system, ts))

def db_fetch_tickets_since(ts_iso: Optional[str]) -> List[Dict[str, Any]]:
    query = """
        SELECT ticket_id, initial_group, group_name AS "group", category,
               author_id, author_name, executor_id, executor_name,
               created_ts, queued_ts, accepted_ts, rejected_ts, closed_ts,
               final_status, reject_reason_code, reject_comment, leader_name, rerouted_to_group,
               clarify_question, clarify_requested_ts, clarify_answer, clarify_answered_ts, updated_ts
        FROM tickets
        WHERE (? IS NULL OR updated_ts > ?)
        ORDER BY updated_ts ASC
    """
    with db() as conn:
        cur = conn.execute(query, (ts_iso, ts_iso))
        return [dict(r) for r in cur.fetchall()]

# ============================================
# РОЛИ / ВЕРИФИКАЦИЯ ПО ТЕЛЕФОНУ
# ============================================

def normalize_phone_e164(raw: str) -> str:
    digits = "".join(ch for ch in (raw or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("8"):
        digits = "7" + digits[1:]
    return "+" + digits

def load_phone_roles_from_env() -> Dict[str, Set[str]]:
    mapping: Dict[str, Set[str]] = {}
    def add_list(env_key: str, role: str):
        raw = os.getenv(env_key, "")
        for item in raw.split(","):
            p = normalize_phone_e164(item.strip())
            if not p:
                continue
            mapping.setdefault(p, set()).add(role)
    add_list("PHONES_AUTHORS", "author")
    add_list("PHONES_EXECUTORS_SVS", "executor:СВС")
    add_list("PHONES_EXECUTORS_SGE", "executor:СГЭ")
    add_list("PHONES_EXECUTORS_SST", "executor:ССТ")
    add_list("PHONES_LEADERS_SVS", "leader:СВС")
    add_list("PHONES_LEADERS_SGE", "leader:СГЭ")
    add_list("PHONES_LEADERS_SST", "leader:ССТ")
    add_list("PHONES_DISPATCHERS", "dispatcher")
    add_list("PHONES_ADMINS", "admin")
    return mapping

PHONE_ROLES_MAP: Dict[str, Set[str]] = {}

def roles_csv(roles: Set[str]) -> str:
    return ",".join(sorted(roles))

def db_roles_ru(user_id: int) -> str:
    roles = db_get_user_roles(user_id)
    if not roles:
        return "нет"
    ru: List[str] = []
    for r in sorted(roles):
        if r == "author":
            ru.append("Заявитель")
        elif r.startswith("executor:"):
            ru.append(f"Исполнитель {r.split(':',1)[1]}")
        elif r.startswith("leader:"):
            ru.append(f"Руководитель {r.split(':',1)[1]}")
        elif r == "dispatcher":
            ru.append("Диспетчер")
        elif r == "admin":
            ru.append("Администратор")
        else:
            ru.append(r)
    return ", ".join(ru)

def has_group_power(user_id: int, group: str) -> bool:
    roles = db_get_user_roles(user_id)
    return (
        f"executor:{group}" in roles
        or f"leader:{group}" in roles
        or "dispatcher" in roles
        or "admin" in roles
    )

def ensure_verified_author(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        u = update.effective_user
        if not u:
            return
        roles = db_get_user_roles(u.id)
        if "author" not in roles and "admin" not in roles and "dispatcher" not in roles:
            await update.message.reply_text(
                "Чтобы отправлять заявки, подтвердите номер телефона в личке с ботом.\nНажмите /verify"
            )
            return
        return await func(update, context)
    return wrapper

# ============================================
# АУДИТ-КАНАЛ
# ============================================

GROUP_TO_ENV = {"СВС": "CHAT_ID_SVS", "СГЭ": "CHAT_ID_SGE", "ССТ": "CHAT_ID_SST"}

def get_group_chat_id(group: str) -> Optional[int]:
    env_key = GROUP_TO_ENV.get(group)
    if not env_key:
        logger.error(f"Unknown group '{group}' (no env key)")
        return None
    val = os.getenv(env_key, "").strip()
    if not val:
        logger.error(f"Env {env_key} not set (group={group})")
        return None
    try:
        return int(val)
    except ValueError:
        logger.error(f"Env {env_key} must be integer chat id, got '{val}'")
        return None

def get_audit_chat_id() -> Optional[int]:
    val = os.getenv("AUDIT_CHAT_ID", "").strip()
    if not val:
        return None
    try:
        return int(val)
    except ValueError:
        logger.error(f"AUDIT_CHAT_ID must be integer chat id, got '{val}'")
        return None

async def audit_log(bot, text: str) -> None:
    chat_id = get_audit_chat_id()
    if chat_id is None:
        return
    try:
        await bot.send_message(chat_id=chat_id, text=text, parse_mode="HTML", disable_web_page_preview=True)
    except Exception:
        logger.exception("audit_log failed")

# ============================================
# ЭВРИСТИКИ (боевые)
# ============================================

GROUP_CATEGORIES: Dict[str, List[Dict[str, Any]]] = {
    "СВС": [
        {"title": "Настройка температуры / обдува",
         "kw": ["обдув", "кондиционер", "тепло", "холодно", "температура в помещении", "душно", "температурный режим",
                "температур", "режим", "прохлад"]},
        {"title": "Ремонт вентиляции / кондиционера",
         "kw": ["вентиляция", "кондиционер", "холодно", "душно", "жарко", "температура", "режим",
                "вентиляц", "температур"]},
        {"title": "Протечки",
         "kw": ["протечка", "капает", "потолок", "вода", "туалет", "раковина", "умывальник",
                "протеч", "капл"]},
        {"title": "Автоматическая установка пожаротушения",
         "kw": ["автоматическая установка пожаротушения", "проверка оборудования", "установка оборудования",
                "пожарный кран", "аупт", "пожаротушен", "спринклер", "ороситель"]},
        {"title": "Засор",
         "kw": ["засор", "туалет", "раковина", "умывальник", "вода", "сантехник", "засорение",
                "канализац", "забилось", "не уходит вода", "слив", "трап"]},
        {"title": "Неприятный запах",
         "kw": ["запах", "туалет", "воняет", "вентиляция", "канализация", "амбре", "смрад"]},
        {"title": "Отопление",
         "kw": ["отопление", "тепло", "холодно", "радиатор", "батаре", "котельн"]},
    ],
    "СГЭ": [
        {"title": "Замена освещения",
         "kw": ["лампочка", "освещение", "свет", "перегорела", "заменить лампу", "замена света"]},
        {"title": "Неисправность / монтаж розетки",
         "kw": ["розетка", "электричество", "контакт", "искрит", "штепсель", "монтаж розетки"]},
        {"title": "Электрощит / питание",
         "kw": ["электрощит", "питание", "свет", "щиток", "автомат", "фаза", "выбивает"]},
        {"title": "Эвакуационное освещение",
         "kw": ["освещение", "эвакуация", "свет", "эвакуацион", "exit"]},
        {"title": "Провода",
         "kw": ["провода", "оголенные", "опасность", "поражение электрическим током", "кабель", "проводка", "оголен", "оголён"]},
        {"title": "Направление освещения",
         "kw": ["освещение", "перенаправить", "свет", "лампочка", "свет на этаже", "направлен", "угол света", "перенастроить свет"]},
    ],
    "ССТ": [
        {"title": "Выключить / включить музыку",
         "kw": ["музыка", "громкость", "играет не громко", "музыкальное сопровождение", "включить музыку", "выключить музыку", "звук"]},
        {"title": "Датчики дымовые (отключение, демонтаж/монтаж)",
         "kw": ["датчик", "дым", "датчик дыма", "дымовой", "пожарный датчик", "отключить датчик", "демонтаж"]},
        {"title": "Оповещатель речевой (отключение, демонтаж/монтаж)",
         "kw": ["речевой оповещатель", "оповещатель", "громкоговор", "сирена"]},
        {"title": "Настройка/ремонт систем автоматики",
         "kw": ["система автоматики", "настройка системы", "ремонт автоматики", "проверка автоматики", "автоматика", "контроллер", "система управлен"]},
    ],
}

def _count_hits(text: str, patterns: List[str]) -> int:
    t = (text or "").lower()
    return sum(1 for p in patterns if p.lower() in t)

def _score_category(text: str, title: str, kw: List[str]) -> int:
    score = _count_hits(text, kw)
    if fuzz.partial_ratio(title.lower(), (text or "").lower()) >= 85:
        score += 1
    return score

def classify(text: str) -> Dict[str, Any]:
    best_group, best_category, best_score = "Неопределено", "Другое", 0
    hits: Dict[str, int] = {}
    for group, cats in GROUP_CATEGORIES.items():
        for c in cats:
            title, kw = c["title"], c["kw"]
            s = _score_category(text, title, kw)
            hits[f"{group}:{title}"] = s
            if s > best_score:
                best_score, best_group, best_category = s, group, title
    return {"group": best_group, "category": best_category, "confidence": 0.0, "hits": hits}

# ============================================
# РУССКИЕ СТАТУСЫ (для UI)
# ============================================

STATUS_RU = {
    "created": "новая",
    "queued": "на отправке",
    "accepted": "в работе",
    "rejected": "отклонена",
    "closed": "закрыта",
    "clarifying": "на уточнении",
}
def status_ru(code: str) -> str:
    return STATUS_RU.get(code, code or "-")

# ============================================
# РЕНДЕР/КНОПКИ
# ============================================

def user_link_html(user_id: int, name: str | None) -> str:
    safe = html_escape(name or "пользователь", quote=False)
    return f'<a href="tg://user?id={user_id}">{safe}</a>'

def ticket_group_text(t: Dict[str, Any]) -> str:
    submit_link = user_link_html(t["submitter_id"], t.get("submitter_name"))
    body = html_escape((t.get("text") or "").strip(), quote=False)
    parts = [
        f"🆕 Заявка #{t['id']} (группа: {t['classification']['group']} / категория: {t['classification']['category']})",
        f"Автор: {submit_link}",
        "",
        body,
        "",
        f"Статус: <b>{html_escape(status_ru(t['status']).upper(), quote=False)}</b>",
    ]
    if t.get("executor_id"):
        parts.append(f"Исполнитель: {user_link_html(t['executor_id'], t.get('executor_name'))}")
    if t.get("reject_reason_code"):
        reasons_ru = {"not_uto":"Не к УТО","other_group":"К другой группе","no_access":"Нет доступа к помещению"}
        parts.append(f"Причина отклонения: {reasons_ru.get(t['reject_reason_code'], t['reject_reason_code'])}")
    if t.get("reject_comment"):
        parts.append(f"Комментарий: {html_escape(t['reject_comment'], quote=False)}")
    if t.get("clarify_question") and t["status"] == "clarifying":
        parts.append(f"🔎 На уточнении: {html_escape(t['clarify_question'], False)}")
    if t.get("pending_reject"):
        parts.append("⏳ Отклонение на согласовании у руководителя.")
    return "\n".join(parts)

def kb_initial(ticket_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Принять", callback_data=f"t:accept:{ticket_id}")],
            [InlineKeyboardButton("⛔ Отклонить", callback_data=f"t:reject:{ticket_id}")],
            [InlineKeyboardButton("🔎 Уточнить", callback_data=f"t:clarify:{ticket_id}")],
        ]
    )

def kb_after_accept(ticket_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Завершить", callback_data=f"t:complete:{ticket_id}")],
            [InlineKeyboardButton("⛔ Отклонить", callback_data=f"t:reject:{ticket_id}")],
            [InlineKeyboardButton("🔎 Уточнить", callback_data=f"t:clarify:{ticket_id}")],
        ]
    )

def kb_reject_reasons(ticket_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("1) Не к УТО", callback_data=f"t:rejchoose:{ticket_id}:not_uto")],
            [InlineKeyboardButton("2) К другой группе", callback_data=f"t:rejchoose:{ticket_id}:other_group")],
            [InlineKeyboardButton("3) Нет доступа к помещению", callback_data=f"t:rejchoose:{ticket_id}:no_access")],
        ]
    )

def kb_leader_approve_or_cancel(ticket_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Согласовать отклонение", callback_data=f"t:leadapprove:{ticket_id}")],
            [InlineKeyboardButton("↩️ Отменить отклонение", callback_data=f"t:leadcancel:{ticket_id}")],
        ]
    )

def kb_leader_choose_group(ticket_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("СВС", callback_data=f"t:leadroute:{ticket_id}:СВС")],
            [InlineKeyboardButton("СГЭ", callback_data=f"t:leadroute:{ticket_id}:СГЭ")],
            [InlineKeyboardButton("ССТ", callback_data=f"t:leadroute:{ticket_id}:ССТ")],
            [InlineKeyboardButton("↩️ Отменить отклонение", callback_data=f"t:leadcancel:{ticket_id}")],
        ]
    )

def main_menu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("📱 Подтвердить номер", callback_data="ui:verify")],
            [InlineKeyboardButton("📊 Экспорт Excel", callback_data="ui:export_excel"),
             InlineKeyboardButton("🧾 Экспорт CSV", callback_data="ui:export_csv")],
            [InlineKeyboardButton("ℹ️ Помощь", callback_data="ui:help")],
        ]
    )

def verify_reply_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [[KeyboardButton("📱 Подтвердить номер (отправить контакт)", request_contact=True)]],
        resize_keyboard=True, one_time_keyboard=True
    )

# ============================================
# СЛУЖЕБНЫЕ СТРУКТУРЫ (ожидания реплая)
# ============================================

# Ключ: (chat_id, message_id) приглашения → данные ожидания
REPLY_WAIT: Dict[Tuple[int, int], Dict[str, Any]] = {}

# Для уточнений: авторский реплай
CLARIFY_AUTHOR_WAIT: Dict[Tuple[int, int], Dict[str, Any]] = {}

# Живые заявки
TICKETS: Dict[str, Dict[str, Any]] = {}

# ============================================
# ОТПРАВКА В ЧАТ ГРУППЫ / РУКОВОДИТЕЛЮ
# ============================================

GROUP_TO_ENV = {"СВС": "CHAT_ID_SVS", "СГЭ": "CHAT_ID_SGE", "ССТ": "CHAT_ID_SST"}

def get_group_chat_id(group: str) -> Optional[int]:
    env_key = GROUP_TO_ENV.get(group)
    val = os.getenv(env_key, "").strip() if env_key else ""
    try:
        return int(val) if val else None
    except Exception:
        logger.error(f"Bad chat id in ENV for {group}: {val}")
        return None

def get_audit_chat_id() -> Optional[int]:
    val = os.getenv("AUDIT_CHAT_ID", "").strip()
    try:
        return int(val) if val else None
    except Exception:
        logger.error(f"Bad AUDIT_CHAT_ID: {val}")
        return None

async def audit_log(bot, text: str) -> None:
    chat_id = get_audit_chat_id()
    if not chat_id:
        return
    try:
        await bot.send_message(chat_id=chat_id, text=text, parse_mode="HTML", disable_web_page_preview=True)
    except Exception:
        logger.exception("audit_log failed")

def user_link_html(uid: int, name: Optional[str]) -> str:
    safe = html_escape(name or "пользователь", quote=False)
    return f'<a href="tg://user?id={uid}">{safe}</a>'

def ticket_group_text(t: Dict[str, Any]) -> str:
    submit_link = user_link_html(t["submitter_id"], t.get("submitter_name"))
    body = html_escape((t.get("text") or "").strip(), quote=False)
    parts = [
        f"🆕 Заявка #{t['id']} (группа: {t['classification']['group']} / категория: {t['classification']['category']})",
        f"Автор: {submit_link}",
        "",
        body,
        "",
        f"Статус: <b>{html_escape(status_ru(t['status']).upper(), quote=False)}</b>",
    ]
    if t.get("executor_id"):
        parts.append(f"Исполнитель: {user_link_html(t['executor_id'], t.get('executor_name'))}")
    if t.get("reject_reason_code"):
        reasons_ru = {"not_uto":"Не к УТО","other_group":"К другой группе","no_access":"Нет доступа к помещению"}
        parts.append(f"Причина отклонения: {reasons_ru.get(t['reject_reason_code'], t['reject_reason_code'])}")
    if t.get("reject_comment"):
        parts.append(f"Комментарий: {html_escape(t['reject_comment'], quote=False)}")
    if t.get("clarify_question") and t["status"] == "clarifying":
        parts.append(f"🔎 На уточнении: {html_escape(t['clarify_question'], False)}")
    if t.get("pending_reject"):
        parts.append("⏳ Отклонение на согласовании у руководителя.")
    return "\n".join(parts)

async def send_to_group(bot, t: Dict[str, Any]) -> Optional[Message]:
    group = t["classification"]["group"]
    chat_id = get_group_chat_id(group)
    if not chat_id:
        return None
    kb = kb_after_accept(t["id"]) if t.get("status") == "accepted" else kb_initial(t["id"])
    try:
        msg = await bot.send_message(chat_id=chat_id, text=ticket_group_text(t), reply_markup=kb, parse_mode="HTML")
        return msg
    except Exception:
        logger.exception("send_to_group failed")
        return None

def db_find_users_by_role_prefix(prefix: str) -> List[sqlite3.Row]:
    with db() as conn:
        cur = conn.execute("SELECT * FROM users WHERE active=1 AND roles LIKE ?", (f"%{prefix}%",))
        return cur.fetchall()

async def send_to_leaders(bot, group: str, text: str, kb: InlineKeyboardMarkup) -> List[int]:
    leaders = db_find_users_by_role_prefix(f"leader:{group}")
    sent: List[int] = []
    for r in leaders:
        leader_id = r["telegram_user_id"]
        try:
            await bot.send_message(chat_id=leader_id, text=text, reply_markup=kb, parse_mode="HTML")
            sent.append(leader_id)
        except Exception:
            logger.exception("send_to_leaders failed")
    return sent

# ============================================
# КОМАНДЫ
# ============================================

def get_admins() -> set[int]:
    return {int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

def admin_only(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not update.effective_user or update.effective_user.id not in get_admins():
            return await update.message.reply_text("Только для админов.")
        return await func(update, context)
    return wrapper

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type == "private":
        u = update.effective_user
        await update.message.reply_text(
            "Привет! Я помогаю отправлять заявки в УТО (СВС/СГЭ/ССТ).\n"
            "1) Подтвердите номер — /verify.\n"
            "2) Напишите текст заявки — и подтвердите отправку.\n"
            f"Ваши роли: {db_roles_ru(u.id) if u else '—'}",
            reply_markup=main_menu_kb()
        )
        await update.message.reply_text("Нажмите кнопку ниже, чтобы отправить боту ваш номер:", reply_markup=verify_reply_kb())
    else:
        await update.message.reply_text("Бот активен. Для помощи — /help или /panel")

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Команды:\n"
        "/start — приветствие\n"
        "/menu, /panel — показать панель кнопок\n"
        "/verify — подтвердить номер телефона\n"
        "/whoami — показать ваш user_id и chat_id\n"
        "/echo_chat_id_any — chat_id текущего чата (диагностика)\n"
        "/echo_chat_id — то же, но только для админов\n"
        "/debug_env — показать chat_id групп и аудит-канала (админ)\n"
        "/export_excel — выгрузить Excel\n"
        "/export_csv — выгрузить CSV\n\n"
        "Важно: когда бот просит комментарий — отвечайте РЕПЛАЕМ на сообщение бота."
    )

async def menu_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Главное меню:", reply_markup=main_menu_kb())

async def panel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Панель:", reply_markup=main_menu_kb())

@admin_only
async def echo_chat_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"chat_id: {update.effective_chat.id}")

async def echo_chat_id_any(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"chat_id: {update.effective_chat.id}")

async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u, c = update.effective_user, update.effective_chat
    await update.message.reply_text(
        f"user_id: {u.id if u else 'unknown'}\n"
        f"chat_id: {c.id}\n"
        f"chat_type: {c.type}\n"
        f"Роли: {db_roles_ru(u.id) if u else '—'}"
    )

@admin_only
async def debug_env(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        f"CHAT_ID_SVS={get_group_chat_id('СВС')}\n"
        f"CHAT_ID_SGE={get_group_chat_id('СГЭ')}\n"
        f"CHAT_ID_SST={get_group_chat_id('ССТ')}\n"
        f"AUDIT_CHAT_ID={get_audit_chat_id()}"
    )

async def verify_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Нажмите кнопку ниже, чтобы отправить боту ваш номер телефона:", reply_markup=verify_reply_kb())

# ============================================
# ВЕРИФИКАЦИЯ КОНТАКТА
# ============================================

def roles_csv(roles: Set[str]) -> str:
    return ",".join(sorted(roles))

PHONE_ROLES_MAP: Dict[str, Set[str]] = {}

def load_phone_roles_from_env() -> Dict[str, Set[str]]:
    mapping: Dict[str, Set[str]] = {}
    def add_list(env_key: str, role: str):
        raw = os.getenv(env_key, "")
        for item in raw.split(","):
            p = normalize_phone_e164(item.strip())
            if not p:
                continue
            mapping.setdefault(p, set()).add(role)
    add_list("PHONES_AUTHORS", "author")
    add_list("PHONES_EXECUTORS_SVS", "executor:СВС")
    add_list("PHONES_EXECUTORS_SGE", "executor:СГЭ")
    add_list("PHONES_EXECUTORS_SST", "executor:ССТ")
    add_list("PHONES_LEADERS_SVS", "leader:СВС")
    add_list("PHONES_LEADERS_SGE", "leader:СГЭ")
    add_list("PHONES_LEADERS_SST", "leader:ССТ")
    add_list("PHONES_DISPATCHERS", "dispatcher")
    add_list("PHONES_ADMINS", "admin")
    return mapping

def normalize_phone_e164(raw: str) -> str:
    digits = "".join(ch for ch in (raw or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("8"):
        digits = "7" + digits[1:]
    return "+" + digits

def db_upsert_user(telegram_user_id: int, phone: str, full_name: str, roles_csv_str: str) -> None:
    with db() as conn:
        conn.execute("""
            INSERT INTO users(telegram_user_id, phone_e164, full_name, roles, verified_at, active)
            VALUES (?, ?, ?, ?, ?, 1)
            ON CONFLICT(telegram_user_id) DO UPDATE SET
                phone_e164=excluded.phone_e164,
                full_name=excluded.full_name,
                roles=excluded.roles,
                verified_at=excluded.verified_at,
                active=1
        """, (telegram_user_id, phone, full_name, roles_csv_str, iso_now()))

def db_get_user_roles(telegram_user_id: int) -> Set[str]:
    with db() as conn:
        r = conn.execute("SELECT roles, active FROM users WHERE telegram_user_id=?", (telegram_user_id,)).fetchone()
        if not r or r["active"] != 1:
            return set()
        roles_csv_str = r["roles"] or ""
        return {x.strip() for x in roles_csv_str.split(",") if x.strip()}

async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.contact:
        return
    c = update.message.contact
    u = update.effective_user
    phone = normalize_phone_e164(c.phone_number)
    roles = PHONE_ROLES_MAP.get(phone, set())
    if not roles:
        await update.message.reply_text(f"Номер {phone} не найден в списке доступа. Обратитесь к администратору.")
        return
    db_upsert_user(u.id, phone, u.full_name or "", roles_csv(roles))
    await update.message.reply_text(
        f"Готово! Номер подтверждён: {phone}\n"
        f"Ваши роли: {db_roles_ru(u.id)}",
        reply_markup=ReplyKeyboardRemove()
    )

# ============================================
# ПРИЁМ ТЕКСТА (новая заявка ИЛИ реплай-комментарий)
# ============================================

def db_roles_ru(user_id: int) -> str:
    roles = db_get_user_roles(user_id)
    if not roles:
        return "нет"
    ru: List[str] = []
    for r in sorted(roles):
        if r == "author":
            ru.append("Заявитель")
        elif r.startswith("executor:"):
            ru.append(f"Исполнитель {r.split(':',1)[1]}")
        elif r.startswith("leader:"):
            ru.append(f"Руководитель {r.split(':',1)[1]}")
        elif r == "dispatcher":
            ru.append("Диспетчер")
        elif r == "admin":
            ru.append("Администратор")
        else:
            ru.append(r)
    return ", ".join(ru)

@ensure_verified_author
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return

    # Если это реплай на приглашение — разбираем в спец обработчике
    if update.message.reply_to_message:
        await handle_reply(update, context)
        return

    text = update.message.text or ""
    u = update.effective_user
    ch = update.effective_chat

    try:
        result = classify(text)
        group = result.get("group", "Неопределено")
        category = result.get("category", "Другое")

        t_id = uuid.uuid4().hex[:8].upper()
        ticket = {
            "id": t_id,
            "submitter_id": u.id if u else None,
            "submitter_name": (u.full_name if u else None),
            "submitter_chat_id": ch.id if ch else None,
            "text": text,
            "classification": result,
            "status": "created",
            "created_ts": iso_now(),
            "initial_group": None,   # проставим при первом QUEUED
        }
        context.user_data["last_ticket"] = ticket

        event = {"event": "new_text", **ticket}
        save_ticket_event_jsonl(event)
        db_insert_event(event)
        db_upsert_ticket_snapshot(ticket)
        db_touch_ticket_timestamp(t_id, "created_ts")

        kb = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("Подтвердить отправку в группу", callback_data="ticket_confirm")],
                [InlineKeyboardButton("Сообщить об ошибке", callback_data="ticket_report_mistake")],
            ]
        )
        msg = (
            "Предварительная классификация (УТО):\n"
            f"• Группа-исполнитель: <b>{group}</b>\n"
            f"• Категория: <b>{category}</b>\n"
            f"• Номер заявки: <b>#{t_id}</b>\n\n"
            "Если всё верно — подтвердите отправку в группу."
        )
        await update.message.reply_html(msg, reply_markup=kb)

        await audit_log(context.bot, f"📝 <b>Draft ticket</b> #{t_id} from {user_link_html(ticket['submitter_id'], ticket['submitter_name'])}\n"
                                     f"Group: {group} / Category: {category}")

    except Exception:
        logger.exception("handle_text failed")
        await update.message.reply_text("Ошибка обработки заявки. Попробуйте ещё раз или /help.")

# ============================================
# EXPORT (из SQLite)
# ============================================

def _dur_str(delta: Optional[timedelta]) -> str:
    if not delta:
        return ""
    total_sec = int(delta.total_seconds())
    h, rem = divmod(total_sec, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"

def _parse_iso(ts: Optional[str]) -> Optional[datetime]:
    if not ts:
        return None
    try:
        if ts.endswith("Z"):
            ts = ts[:-1]
        dt = datetime.fromisoformat(ts)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt.astimezone(UTC)
    except Exception:
        return None

def aggregate_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for r in rows:
        created = _parse_iso(r.get("created_ts"))
        queued  = _parse_iso(r.get("queued_ts"))
        accepted= _parse_iso(r.get("accepted_ts"))
        rejected= _parse_iso(r.get("rejected_ts"))
        closed  = _parse_iso(r.get("closed_ts"))
        clarify_req = _parse_iso(r.get("clarify_requested_ts"))
        clarify_ans = _parse_iso(r.get("clarify_answered_ts"))

        r["time_to_queue"]   = _dur_str((queued - created) if (created and queued) else None)
        r["time_to_accept"]  = _dur_str((accepted - queued) if (accepted and queued) else None)
        r["time_in_progress"]= _dur_str((closed - accepted) if (closed and accepted) else None)
        r["time_to_clarify"] = _dur_str((clarify_ans - clarify_req) if (clarify_req and clarify_ans) else None)

        end_ts = closed or rejected
        r["total_time"]      = _dur_str((end_ts - created) if (end_ts and created) else None)
    return rows

def _to_excel_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    try:
        return json.dumps(v, ensure_ascii=False)
    except Exception:
        return str(v)

def write_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    header = [
        "ticket_id","initial_group","group","category","author_id","author_name","executor_id","executor_name",
        "created_ts","queued_ts","accepted_ts","rejected_ts","closed_ts",
        "time_to_queue","time_to_accept","time_in_progress","time_to_clarify","total_time",
        "final_status","reject_reason_code","reject_comment","leader_name","rerouted_to_group",
        "clarify_question","clarify_requested_ts","clarify_answer","clarify_answered_ts"
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(header)
        for r in rows:
            writer.writerow([r.get(k, "") for k in header])

def write_xlsx(rows: List[Dict[str, Any]], path: Path) -> Tuple[bool, str]:
    try:
        from openpyxl import Workbook
    except Exception:
        return False, "openpyxl не установлен"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "tickets"
    header = [
        "ticket_id","initial_group","group","category","author_id","author_name","executor_id","executor_name",
        "created_ts","queued_ts","accepted_ts","rejected_ts","closed_ts",
        "time_to_queue","time_to_accept","time_in_progress","time_to_clarify","total_time",
        "final_status","reject_reason_code","reject_comment","leader_name","rerouted_to_group",
        "clarify_question","clarify_requested_ts","clarify_answer","clarify_answered_ts"
    ]
    ws1.append(header)
    for r in rows:
        ws1.append([_to_excel_cell(r.get(k, "")) for k in header])
    # авто-ширина
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(header)+1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws1[letter]:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws1.column_dimensions[letter].width = min(max(10, max_len + 2), 60)
    for i in range(3):
        try:
            wb.save(path)
            return True, "ok"
        except Exception as ex:
            if i == 2:
                return False, f"save failed: {ex}"
            time.sleep(0.5)

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rows = aggregate_rows(db_fetch_tickets_rows())
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"tickets_{ts}.xlsx"
    ok, msg = write_xlsx(rows, out_path)
    if not ok:
        await update.message.reply_text(f"Не удалось сформировать Excel: {msg}. Попробуйте /export_csv.")
        return
    with out_path.open("rb") as f:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=InputFile(f, filename=out_path.name),
            caption=f"Экспорт заявок (актуально на {ts} UTC)."
        )
    await audit_log(context.bot, f"📊 Export Excel sent ({out_path.name})")

async def export_csv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rows = aggregate_rows(db_fetch_tickets_rows())
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    out_path = DATA_DIR / f"tickets_{ts}.csv"
    write_csv(rows, out_path)
    with out_path.open("rb") as f:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=InputFile(f, filename=out_path.name),
            caption=f"Экспорт заявок (CSV, {ts} UTC)."
        )
    await audit_log(context.bot, f"📊 Export CSV sent ({out_path.name})")

# ============================================
# CALLBACKS: подтверждение/исполнитель/руководитель
# ============================================

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query:
        return
    data = query.data or ""
    user = update.effective_user

    try:
        # UI
        if data == "ui:help":
            await help_cmd(update, context); await query.answer(); return
        if data == "ui:export_excel":
            await export_excel(update, context); await query.answer(); return
        if data == "ui:export_csv":
            await export_csv(update, context); await query.answer(); return
        if data == "ui:verify":
            if update.effective_chat.type == "private":
                await verify_cmd(update, context)
            else:
                await query.message.reply_text("Подтверждение номера доступно только в личке с ботом: откройте диалог и нажмите /verify")
            await query.answer(); return

        if data == "ticket_confirm":
            ticket = context.user_data.get("last_ticket")
            if not ticket:
                await query.answer("Не найден контекст заявки, отправьте текст ещё раз.")
                return

            # фиксируем initial_group при первой отправке
            if not ticket.get("initial_group"):
                ticket["initial_group"] = ticket["classification"]["group"]

            ticket["status"] = "queued"
            TICKETS[ticket["id"]] = ticket

            event = {"event": "queued_to_group", **ticket}
            save_ticket_event_jsonl(event)
            db_insert_event(event)
            db_upsert_ticket_snapshot(ticket)
            db_touch_ticket_timestamp(ticket["id"], "queued_ts")

            msg = await send_to_group(context.bot, ticket)
            if msg:
                ticket["group_chat_id"] = msg.chat.id
                ticket["group_message_id"] = msg.message_id
                db_upsert_ticket_snapshot(ticket)
                await query.answer("Заявка отправлена в группу.")
                await query.edit_message_reply_markup(reply_markup=None)
                await audit_log(context.bot, f"📤 Sent to group #{ticket['id']} → {ticket['classification']['group']} / {ticket['classification']['category']}")
            else:
                await query.answer("Не удалось отправить в чат группы. Проверьте настройки.")
                try:
                    await context.bot.send_message(
                        chat_id=ticket["submitter_chat_id"],
                        text="Не удалось отправить вашу заявку в чат группы. Обратитесь к администратору."
                    )
                except Exception:
                    pass
            return

        if data == "ticket_report_mistake":
            save_feedback_jsonl({"user_id": user.id if user else None, "feedback": "heuristics_mistake"})
            await query.answer("Принято. Улучшим правила.")
            try: await query.edit_message_reply_markup(reply_markup=None)
            except Exception: pass
            await audit_log(context.bot, f"⚠️ Heuristics mistake reported by user_id={user.id if user else 'unknown'}")
            return

        if not data.startswith("t:"):
            await query.answer("Неизвестная команда.")
            return

        parts = data.split(":")
        action = parts[1]
        t_id = parts[2] if len(parts) > 2 else None
        t = TICKETS.get(t_id) if t_id else None

        # Проверка прав по группе для действий исполнителя
        if action in {"accept","reject","clarify","complete"}:
            if not t:
                await query.answer("Заявка не найдена (возможно, бот перезапускался).")
                return
            group = t["classification"]["group"]
            if not has_group_power(user.id, group):
                await query.answer("Недостаточно прав для действий по этой заявке.", show_alert=True)
                return
            if query.message and (t.get("group_chat_id") != query.message.chat.id or t.get("group_message_id") != query.message.message_id):
                await query.answer("Это сообщение уже устарело.")
                return

        if action == "accept":
            if t["status"] in {"accepted", "closed"}:
                await query.answer("Уже в работе/закрыта.")
                return
            t["status"] = "accepted"
            t["executor_id"] = user.id
            t["executor_name"] = user.full_name
            TICKETS[t_id] = t

            save_ticket_event_jsonl({"event": "accepted", "ticket_id": t_id, "executor_id": user.id})
            db_insert_event({"event": "accepted", "ticket_id": t_id, "executor_id": user.id,
                             "group": group, "category": t["classification"]["category"]})
            db_upsert_ticket_snapshot(t)
            db_touch_ticket_timestamp(t_id, "accepted_ts")

            try:
                await query.edit_message_text(text=ticket_group_text(t), reply_markup=kb_after_accept(t_id), parse_mode="HTML")
            except Exception:
                logger.exception("edit after accept failed")

            try:
                await context.bot.send_message(
                    chat_id=t["submitter_chat_id"],
                    text=(f"Заявка #{t_id} принята в работу.\nИсполнитель: {user_link_html(user.id, user.full_name)}"),
                    parse_mode="HTML",
                )
            except Exception:
                logger.exception("notify submitter accept failed")

            await audit_log(context.bot, f"✅ Accepted #{t_id} by {user_link_html(user.id, user.full_name)}")
            await query.answer("Взято в работу.")
            return

        if action == "reject":
            await query.answer()
            prompt = await query.message.reply_text(
                "Выберите причину отклонения ниже, затем напишите комментарий РЕПЛАЕМ на это сообщение.",
                reply_markup=kb_reject_reasons(t_id)
            )
            REPLY_WAIT[(prompt.chat.id, prompt.message_id)] = {
                "type": "reject_comment_wait",
                "ticket_id": t_id,
                "executor_id": user.id,
                "reason_code": None,
            }
            return

        if action == "rejchoose":
            reason_code = parts[3] if len(parts) > 3 else None
            if reason_code not in {"not_uto","other_group","no_access"}:
                await query.answer("Неизвестная причина.")
                return
            found_key = None
            for k, v in list(REPLY_WAIT.items()):
                if v.get("type") == "reject_comment_wait" and v.get("ticket_id") == t_id and v.get("executor_id") == user.id:
                    found_key = k
            if found_key is None:
                prompt = await query.message.reply_text("Напишите комментарий РЕПЛАЕМ на это сообщение (почему отклоняете).")
                found_key = (prompt.chat.id, prompt.message_id)
                REPLY_WAIT[found_key] = {
                    "type": "reject_comment_wait",
                    "ticket_id": t_id,
                    "executor_id": user.id,
                    "reason_code": reason_code,
                }
            else:
                REPLY_WAIT[found_key]["reason_code"] = reason_code

            await query.answer("Причина зафиксирована. Введите комментарий РЕПЛАЕМ.")
            return

        if action == "clarify":
            await query.answer()
            prompt = await query.message.reply_text("Введите уточняющий вопрос РЕПЛАЕМ на это сообщение — мы отправим его автору.")
            REPLY_WAIT[(prompt.chat.id, prompt.message_id)] = {
                "type": "clarify_question",
                "ticket_id": t_id,
                "executor_id": user.id,
            }
            return

        if action == "complete":
            if t["status"] != "accepted":
                await query.answer("Сначала возьмите заявку в работу.")
                return
            roles = db_get_user_roles(user.id)
            if (t.get("executor_id") not in (None, user.id)) and (f"leader:{t['classification']['group']}" not in roles) and ("admin" not in roles):
                await query.answer("Завершить может только принявший исполнитель (или руководитель/админ).")
                return

            t["status"] = "closed"
            t["closed"] = True
            t["completed_by"] = user.id
            TICKETS[t_id] = t

            save_ticket_event_jsonl({"event": "closed_by_executor", "ticket_id": t_id, "executor_id": user.id})
            db_insert_event({"event": "closed_by_executor", "ticket_id": t_id, "executor_id": user.id,
                             "group": t["classification"]["group"], "category": t["classification"]["category"]})
            db_upsert_ticket_snapshot(t)
            db_touch_ticket_timestamp(t_id, "closed_ts")

            try:
                await query.edit_message_text(text=ticket_group_text(t), reply_markup=None, parse_mode="HTML")
            except Exception:
                logger.exception("edit after close failed")

            try:
                await context.bot.send_message(
                    chat_id=t["submitter_chat_id"],
                    text=(f"Исполнитель {user_link_html(user.id, user.full_name)} закрыл заявку #{t_id}. ✅"),
                    parse_mode="HTML",
                )
            except Exception:
                logger.exception("notify submitter closed failed")

            await audit_log(context.bot, f"🧾 Closed #{t_id} by {user_link_html(user.id, user.full_name)}")
            await query.answer("Заявка закрыта.")
            return

        # Действия руководителя
        if action in {"leadapprove","leadcancel","leadroute"}:
            if not t:
                await query.answer("Заявка не найдена.")
                return
            group = t["classification"]["group"]
            roles = db_get_user_roles(user.id)
            if not (f"leader:{group}" in roles or "dispatcher" in roles or "admin" in roles):
                await query.answer("Только для руководителя соответствующей группы (или диспетчера/админа).", show_alert=True)
                return

            if action == "leadapprove":
                pend = t.get("pending_reject")
                if not pend or pend.get("reason_code") == "other_group":
                    await query.answer("Нет ожидающего отклонения (или выбрана маршрутизация).")
                    return
                t["status"] = "rejected"
                t["reject_reason_code"] = pend["reason_code"]
                t["reject_comment"] = pend.get("comment")
                t["leader_id"] = user.id
                t["leader_name"] = user.full_name
                t["leader_decision_ts"] = iso_now()
                t["rejected_ts"] = t.get("rejected_ts") or iso_now()
                t.pop("pending_reject", None)
                TICKETS[t_id] = t

                save_ticket_event_jsonl({"event": "rejected", "ticket_id": t_id, "executor_id": pend["executor_id"], "leader_id": user.id, "comment": t["reject_comment"]})
                db_insert_event({"event": "rejected", "ticket_id": t_id, "executor_id": pend["executor_id"], "leader_id": user.id,
                                 "group": group, "category": t["classification"]["category"], "comment": t["reject_comment"]})
                db_upsert_ticket_snapshot(t)
                db_touch_ticket_timestamp(t_id, "rejected_ts")

                try:
                    await context.bot.edit_message_text(
                        chat_id=t["group_chat_id"],
                        message_id=t["group_message_id"],
                        text=ticket_group_text(t),
                        parse_mode="HTML",
                        reply_markup=None,
                    )
                except Exception:
                    logger.exception("edit group after leader approve failed")

                try:
                    await context.bot.send_message(
                        chat_id=t["submitter_chat_id"],
                        text=(f"Заявка #{t_id} отклонена.\n"
                              f"Причина: { {'not_uto':'Не к УТО','no_access':'Нет доступа к помещению'}.get(t['reject_reason_code'],'—') }\n"
                              f"Комментарий: {html_escape(t.get('reject_comment') or '-', False)}"),
                        parse_mode="HTML",
                    )
                except Exception:
                    logger.exception("notify submitter rejected failed")

                await audit_log(context.bot, f"❌ Rejected (leader approved) #{t_id} reason={t['reject_reason_code']}")
                await query.answer("Отклонение согласовано.")
                return

            if action == "leadcancel":
                await query.answer()
                prompt = await context.bot.send_message(
                    chat_id=user.id,
                    text="Отмена отклонения: ответьте РЕПЛАЕМ на это сообщение и укажите комментарий исполнителю (можно пусто)."
                )
                REPLY_WAIT[(prompt.chat.id, prompt.message_id)] = {
                    "type": "leader_cancel_comment",
                    "ticket_id": t_id,
                    "leader_id": user.id,
                }
                return

            if action == "leadroute":
                dest_group = parts[3] if len(parts) > 3 else None
                if dest_group not in {"СВС","СГЭ","ССТ"}:
                    await query.answer("Неизвестная группа.")
                    return
                pend = t.get("pending_reject")
                if not pend or pend.get("reason_code") != "other_group":
                    await query.answer("Маршрутизация не ожидается.")
                    return

                t["classification"]["group"] = dest_group
                t["status"] = "queued"
                t["rerouted_to_group"] = dest_group
                t["rerouted_ts"] = iso_now()
                if not t.get("initial_group"):
                    t["initial_group"] = pend.get("from_group") or group
                t["executor_id"] = None
                t["executor_name"] = None
                t.pop("pending_reject", None)
                TICKETS[t_id] = t

                msg = await send_to_group(context.bot, t)
                if msg:
                    t["group_chat_id"] = msg.chat.id
                    t["group_message_id"] = msg.message_id

                db_upsert_ticket_snapshot(t)
                db_touch_ticket_timestamp(t_id, "queued_ts")
                save_ticket_event_jsonl({"event": "rerouted", "ticket_id": t_id, "leader_id": user.id, "to_group": dest_group})
                db_insert_event({"event": "rerouted", "ticket_id": t_id, "executor_id": pend["executor_id"], "leader_id": user.id,
                                 "group": dest_group, "category": t["classification"]["category"], "to_group": dest_group})

                try:
                    await context.bot.send_message(
                        chat_id=t["submitter_chat_id"],
                        text=(f"Ваша заявка #{t_id} перенаправлена в группу {dest_group}."),
                    )
                except Exception:
                    logger.exception("notify submitter rerouted failed")

                await audit_log(context.bot, f"🔀 Rerouted #{t_id} → {dest_group}")
                await query.answer("Перенаправлено в другую группу.")
                return

        await query.answer("Неизвестное действие.")
    except Exception:
        logger.exception("on_callback failed")
        try:
            await query.answer("Ошибка обработки нажатия.")
        except Exception:
            if query.message:
                await query.message.reply_text("Ошибка обработки нажатия.")

# ============================================
# ОБРАБОТЧИК РЕПЛАЕВ (комментарии/уточнения)
# ============================================

async def handle_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.reply_to_message:
        return
    reply_key = (update.effective_chat.id, update.message.reply_to_message.message_id)
    u = update.effective_user
    text = (update.message.text or "").strip()

    # 1) Комментарий к отклонению от исполнителя
    ctx = REPLY_WAIT.get(reply_key)
    if ctx and ctx.get("type") == "reject_comment_wait":
        t_id = ctx["ticket_id"]
        t = TICKETS.get(t_id)
        if not t:
            REPLY_WAIT.pop(reply_key, None)
            return
        reason = ctx.get("reason_code")
        if reason not in {"not_uto","other_group","no_access"}:
            await update.message.reply_text("Сначала выберите причину отклонения кнопкой, затем повторите комментарий РЕПЛАЕМ.")
            return

        t["pending_reject"] = {
            "executor_id": u.id,
            "executor_name": u.full_name,
            "from_group": t["classification"]["group"],
            "reason_code": reason,
            "comment": text,
            "ts": iso_now(),
        }
        t["reject_reason_code"] = reason
        t["reject_comment"] = text
        TICKETS[t_id] = t
        db_upsert_ticket_snapshot(t)

        try:
            await context.bot.edit_message_text(
                chat_id=t["group_chat_id"],
                message_id=t["group_message_id"],
                text=ticket_group_text(t),
                parse_mode="HTML",
                reply_markup=None,
            )
        except Exception:
            logger.exception("edit group after pending reject failed")

        leader_text = (
            f"⛔ Запрос на отклонение заявки #{t_id}\n"
            f"Группа: {t['classification']['group']} / Категория: {t['classification']['category']}\n"
            f"Исполнитель: {user_link_html(u.id, u.full_name)}\n"
            f"Причина: { {'not_uto':'Не к УТО','other_group':'К другой группе','no_access':'Нет доступа к помещению'}[reason] }\n"
            f"Комментарий: {html_escape(text, False)}\n\n"
            f"{'Выберите группу (для перенаправления):' if reason=='other_group' else 'Доступны действия:'}"
        )
        kb = kb_leader_choose_group(t_id) if reason == "other_group" else kb_leader_approve_or_cancel(t_id)
        leader_ids = await send_to_leaders(context.bot, t["classification"]["group"], leader_text, kb)
        await update.message.reply_text("Отклонение отправлено руководителю на согласование.")
        await audit_log(context.bot, f"⏳ Reject pending #{t_id} reason={reason} (leaders: {leader_ids})")
        REPLY_WAIT.pop(reply_key, None)
        return

    # 2) Вопрос на уточнение от исполнителя
    if ctx and ctx.get("type") == "clarify_question":
        t_id = ctx["ticket_id"]
        t = TICKETS.get(t_id)
        if not t:
            REPLY_WAIT.pop(reply_key, None)
            return

        t["status"] = "clarifying"
        t["clarify_question"] = text
        t["clarify_requested_ts"] = iso_now()
        TICKETS[t_id] = t
        db_upsert_ticket_snapshot(t)

        try:
            await context.bot.edit_message_text(
                chat_id=t["group_chat_id"],
                message_id=t["group_message_id"],
                text=ticket_group_text(t),
                parse_mode="HTML",
                reply_markup=None,
            )
        except Exception:
            logger.exception("edit group after clarify failed")

        try:
            msg = await context.bot.send_message(
                chat_id=t["submitter_chat_id"],
                text=(f"По заявке #{t_id} требуется уточнение от исполнителя "
                      f"{user_link_html(ctx['executor_id'], u.full_name)}:\n\n"
                      f"{html_escape(text, False)}\n\n"
                      f"Пожалуйста, ответьте <b>реплаем на это сообщение</b>."),
                parse_mode="HTML",
            )
            CLARIFY_AUTHOR_WAIT[(t["submitter_chat_id"], msg.message_id)] = {
                "ticket_id": t_id,
                "executor_id": ctx["executor_id"],
            }
        except Exception:
            logger.exception("send clarify to author failed")

        await update.message.reply_text("Вопрос отправлен автору. Ожидаем ответа.")
        await audit_log(context.bot, f"🔎 Clarify requested #{t_id}")
        REPLY_WAIT.pop(reply_key, None)
        return

    # 3) Комментарий руководителя при отмене отклонения
    if ctx and ctx.get("type") == "leader_cancel_comment":
        t_id = ctx["ticket_id"]
        t = TICKETS.get(t_id)
        REPLY_WAIT.pop(reply_key, None)
        if not t:
            return

        ex_id = (t.get("pending_reject") or {}).get("executor_id") or t.get("executor_id")
        if ex_id:
            try:
                await context.bot.send_message(
                    chat_id=ex_id,
                    text=(f"Руководитель отменил отклонение по заявке #{t_id}.\n"
                          f"Комментарий: {html_escape(text or '-', False)}"),
                    parse_mode="HTML",
                )
            except Exception:
                logger.exception("notify executor cancel failed")

        # возвращаем кнопки в групповое сообщение
        try:
            kb = kb_after_accept(t_id) if t.get("executor_id") else kb_initial(t_id)
            await context.bot.edit_message_text(
                chat_id=t["group_chat_id"],
                message_id=t["group_message_id"],
                text=ticket_group_text({**t, "pending_reject": None}),
                parse_mode="HTML",
                reply_markup=kb,
            )
        except Exception:
            logger.exception("edit group after leader cancel failed")

        # закрепляем за изначальным исполнителем
        pend_exec = (t.get("pending_reject") or {}).get("executor_id") or t.get("executor_id")
        if pend_exec:
            t["executor_id"] = pend_exec
            # имя берём из исходной pending, если было
            t["executor_name"] = (t.get("pending_reject") or {}).get("executor_name") or t.get("executor_name")
        t["leader_id"] = u.id
        t["leader_name"] = u.full_name
        t["leader_decision_ts"] = iso_now()
        t.pop("pending_reject", None)
        TICKETS[t_id] = t
        db_upsert_ticket_snapshot(t)
        await audit_log(context.bot, f"↩️ Reject canceled by leader #{t_id}")
        return

    # 4) Ответ автора на уточнение
    info = CLARIFY_AUTHOR_WAIT.get(reply_key)
    if info:
        t_id = info["ticket_id"]
        t = TICKETS.get(t_id)
        CLARIFY_AUTHOR_WAIT.pop(reply_key, None)
        if not t:
            return
        answer = text
        t["clarify_answer"] = answer
        t["clarify_answered_ts"] = iso_now()
        TICKETS[t_id] = t
        db_upsert_ticket_snapshot(t)

        try:
            kb = kb_after_accept(t_id) if (t.get("status") == "accepted" or t.get("executor_id")) else kb_initial(t_id)
            await context.bot.edit_message_text(
                chat_id=t["group_chat_id"],
                message_id=t["group_message_id"],
                text=ticket_group_text({**t, "status": t.get("status") or "queued"}),
                parse_mode="HTML",
                reply_markup=kb,
            )
        except Exception:
            logger.exception("edit group after clarify answer failed")

        try:
            await context.bot.send_message(
                chat_id=info["executor_id"],
                text=(f"Ответ автора по заявке #{t_id}:\n\n{html_escape(answer, False)}"),
                parse_mode="HTML",
            )
        except Exception:
            logger.exception("send clarify answer to executor failed")

        await audit_log(context.bot, f"📩 Clarify answered #{t_id}")
        return

# ============================================
# ОШИБКИ
# ============================================

async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Unhandled error")
    try:
        if update and hasattr(update, "effective_chat") and update.effective_chat:
            await context.bot.send_message(update.effective_chat.id, "Упс. Что-то пошло не так. Попробуйте ещё раз.")
    except TelegramError:
        logger.exception("Failed to notify user about error")

# ============================================
# MAIN
# ============================================

async def on_contact_button_removed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.type == "private":
        try:
            await update.message.reply_text("Готово.", reply_markup=ReplyKeyboardRemove())
        except Exception:
            pass

def main():
    setup_logging(LOGS_DIR)
    load_env(PROJECT_ROOT)
    db_init()
    db_update_from_events()

    global PHONE_ROLES_MAP
    PHONE_ROLES_MAP = load_phone_roles_from_env()

    token = os.getenv("BOT_TOKEN")
    if not token:
        raise RuntimeError("Не найден BOT_TOKEN в .env")

    logger.info(
        f"ENV chat ids: SVS={get_group_chat_id('СВС')} "
        f"SGE={get_group_chat_id('СГЭ')} "
        f"SST={get_group_chat_id('ССТ')} "
        f"AUDIT={get_audit_chat_id()} | DB={DB_PATH}"
    )

    app = ApplicationBuilder().token(token).build()

    # Команды
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", menu_cmd))
    app.add_handler(CommandHandler("panel", panel_cmd))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("verify", verify_cmd))
    app.add_handler(CommandHandler("whoami", whoami))
    app.add_handler(CommandHandler("echo_chat_id_any", echo_chat_id_any))
    app.add_handler(CommandHandler("echo_chat_id", echo_chat_id))  # admin-only
    app.add_handler(CommandHandler("debug_env", debug_env))        # admin-only
    app.add_handler(CommandHandler("export_excel", export_excel))
    app.add_handler(CommandHandler("export_csv", export_csv))

    # Текст от автора (в личке): новая заявка И/ИЛИ реплай-комментарий
    app.add_handler(MessageHandler(filters.ChatType.PRIVATE & filters.TEXT & ~filters.COMMAND, handle_text))

    # Контакт для верификации (в личке)
    app.add_handler(MessageHandler(filters.ChatType.PRIVATE & filters.CONTACT, handle_contact))
    app.add_handler(MessageHandler(filters.ChatType.PRIVATE & filters.Regex("^📱") & ~filters.CONTACT, on_contact_button_removed))

    # Реплаи на приглашения в любом чате
    app.add_handler(MessageHandler(filters.REPLY & filters.TEXT & ~filters.COMMAND, handle_reply))

    # Кнопки
    app.add_handler(CallbackQueryHandler(on_callback))

    app.add_error_handler(on_error)

    logger.info("Bot is starting via long polling...")
    app.run_polling()  # PTB 21.x

if __name__ == "__main__":
    main()
